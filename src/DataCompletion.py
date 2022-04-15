import pandas as pd
import numpy as np
from sklearn.impute import SimpleImputer
import openpyxl as op

data = pd.read_excel(r'excel1.xlsx')
    data_old = pd.read_excel(r'excel1.xlsx')
    data = data.iloc[:,2:-1]
    data_old = data_old.iloc[:, 2:-1]
    column = 0
    for i in range(64):
        for j in range(3717):
            if data.iloc[j, i] != None:
                try:
                    temp = (float)(data.iloc[j, i])
                except ValueError:
                    data.iloc[j, i] = 0
        #isnull = 0
        #for value in data.iloc[:,i].isnull():
            #if value == True: isnull+=1
        #if isnull >= 2910:
            #data.drop(data.columns[column], axis=1, inplace=True)
        #else: column+=1
    sindex = np.argsort(data.isnull().sum()).values
    data.info()
    print(sindex)
    for i in sindex:
        print(i)
        df = data
        fillc = df.iloc[:, i]
        df = df.iloc[:, df.columns != df.columns[i]]
        df_0 = SimpleImputer(missing_values=np.nan
                             , strategy="constant"
                             , fill_value=0
                             ).fit_transform(df)
        Ytrain = fillc[fillc.notnull()]
        Ytest = fillc[fillc.isnull()]

        Xtrain = df_0[Ytrain.index, :]
        Xtest = df_0[Ytest.index, :]

        rfc = RandomForestRegressor()
        rfc.fit(Xtrain, Ytrain)
        Ypredict = rfc.predict(Xtest)

        data.loc[data.iloc[:, i].isnull(), data.columns[i]] = Ypredict
    data.info()
    wb = op.load_workbook("excel1.xlsx")
    sheets = wb.sheetnames
    print(sheets)
    sh = wb[sheets[0]]
    #sh.cell(2,12,0)
    # data_old.info()
    for i in range(64):
        for j in range(3717):
            if pd.isnull(data_old.iloc[j,i]):
                sh.cell(j+2,3+i,data.iloc[j,i])
    wb.save("excel2.xlsx")