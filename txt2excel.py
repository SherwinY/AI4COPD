import openpyxl
import codecs
from openpyxl.utils import get_column_letter


def txt_to_xlsx(filename, outfile):
    fr = codecs.open(filename, 'r', encoding='utf-8')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws = wb.create_sheet()
    ws.title = 'Sheet1'
    row = 0
    for line in fr:
        row += 1
        line = line.strip()
        line = line.split('^')
        col = 0
        for j in range(len(line)):
            col += 1
            ws.cell(column=col, row=row, value=line[j].format(get_column_letter(col)))
    wb.save(outfile)


if __name__ == '__main__':
    inputfileTxt = 'input.txt'
    outfileExcel = 'output.xlsx'
    txt_to_xlsx(inputfileTxt, outfileExcel)
